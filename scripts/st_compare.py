#!/usr/bin/env python
"""Daily ST Files check: our stored numbers vs each property's own
"<Name>-ST" Google Sheet, which is the ground truth for what gets filed.

    .venv/bin/python scripts/st_compare.py             # whatever date the sheets hold
    .venv/bin/python scripts/st_compare.py 2026-08-25  # a specific date

The sheets are NOT a running log - each workbook holds exactly ONE pasted
MEWS "Availability report", for whatever day was last exported into it. So
the date is READ FROM THE SHEETS rather than assumed: with no argument this
compares whatever day they currently hold, which is what makes it a daily
check rather than a one-off. If the eight sheets disagree about the date,
that is reported instead of being silently averaged over.

Reads our side straight from Supabase (st_files_sync) via sync_service, so
the FastAPI server does not need to be running.
"""
import asyncio
import io
import sys
from collections import OrderedDict
from datetime import datetime, timedelta, timezone
from pathlib import Path

import httpx
from openpyxl import load_workbook

# app.config declares env_file=".env" relative to the CWD, so the backend's
# credentials only load when the process is rooted at api/. Chdir before the
# import rather than after - config is read at import time.
import os  # noqa: E402
_API = Path(__file__).resolve().parent.parent / "api"
os.chdir(_API)
sys.path.insert(0, str(_API))
from app.services.sync_service import sync_service  # noqa: E402

# Same IDs as the links the front-office team maintains. The Master tab of
# each is what the filed ST file is built from.
SHEETS = OrderedDict([
    ("Lub d Bangkok Chinatown",       ("Chinatown", "1npf-d74wYYwsQk9LrNUyYJxnCHH0dsLVuxELJo_CqUM")),
    ("Lub d Bangkok Siam",            ("Siam",      "1JFEQcs1lz62KSIYPuzJhY8o_nP_4PEGdZxmeKrqELVc")),
    ("Lub d Koh Samui Chaweng Beach", ("Samui",     "1DwDsPAajjFH5Fbe_-6Jcl-3ds0q-DCkyGVZr2sRpSw4")),
    ("Lub d Koh Tao Tanote Bay",      ("Koh Tao",   "1-48rztEk1J0TmOiFgPfTE9I_UB0Ba9-RvxOKj8aJyG0")),
    ("Lub d Philippines Makati",      ("Makati",    "1ykTLblIv9dIXEDzdivzPr-z6IWDr_asdeCTn0evut3k")),
    ("Lub d Phuket Patong",           ("Patong",    "1X8mh5Hlvcl6hZhO4Z-eyqdZeGQSFwbYjwHTjzMxcEEs")),
    ("Lub d Siem Reap",               ("Siem Reap", "17siV7sMIT5GsW9x8LdsWBVjh1DcSUTchX-WwiS09FpY")),
    ("Marasca Samui",                 ("Marasca",   "1r2i50lPT8VOFKsjqnSgphawAQvgb3S8yWafyGpVIOCs")),
])

# (Master row label, our st_files_sync key). Siam and Siem Reap label the
# last row "Complimentary Room" where the others say "Complimentary".
METRICS = [
    ("Spaces",        "spaces"),
    ("Occupied",      "occupied"),
    ("House uses",    "house_use"),
    ("Out of order",  "out_of_order"),
    ("Availability",  "availability"),
    ("Customers",     "customers"),
    ("Arrivals",      "arrivals"),
    ("Departures",    "departures"),
    ("Complimentary", "complimentary"),
]

# Manila is UTC+8; every other property is UTC+7.
TZ_OFFSET = {"Lub d Philippines Makati": 8}


def read_master(sheet_id: str) -> dict:
    """Master tab as {label: value}, plus the report date and the timestamp
    MEWS generated the underlying export."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    r = httpx.get(url, follow_redirects=True, timeout=90)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), data_only=True)

    master = {}
    ws = wb["Master"]
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()
        if label:
            master[label] = ws.cell(row, 2).value
    # Not every sheet uses the same label for this one.
    if "Complimentary" not in master and "Complimentary Room" in master:
        master["Complimentary"] = master["Complimentary Room"]

    # Prefer the Availability parameters tab; several sheets also carry a
    # "Parameters Reservation"/"Parameters-Reservation" tab for a different
    # export, whose Created time would be the wrong one to report.
    created = None
    params = next((wb[n] for n in ("Parameters", "Parameters-Availability") if n in wb.sheetnames), None)
    if params is None:
        params = next((wb[n] for n in wb.sheetnames
                       if n.startswith("Parameters") and "eservation" not in n), None)
    if params is not None:
        for row in range(1, 21):
            if str(params.cell(row, 1).value or "").strip() == "Created":
                created = params.cell(row, 2).value

    report_date = master.get("Date")
    if isinstance(report_date, datetime):
        report_date = report_date.strftime("%Y-%m-%d")
    return {"master": master, "date": report_date, "created": created}


def as_int(v):
    """Blank Complimentary cells (Siam, Siem Reap) mean zero, not missing."""
    if v is None or v == "":
        return 0
    return int(v) if isinstance(v, (int, float)) else v


def local(ts: str, prop: str) -> str:
    if not ts:
        return "-"
    import re
    ts = re.sub(r"\.(\d+)", lambda m: "." + m.group(1)[:6].ljust(6, "0"), ts)
    off = TZ_OFFSET.get(prop, 7)
    return datetime.fromisoformat(ts).astimezone(timezone(timedelta(hours=off))).strftime("%d %b %H:%M")


async def main():
    want = sys.argv[1] if len(sys.argv) > 1 else None

    print("กำลังโหลดชีตทั้ง 8 ...", flush=True)
    sheets = {}
    for prop, (short, sid) in SHEETS.items():
        try:
            sheets[prop] = read_master(sid)
        except Exception as e:
            print(f"  !! {short}: อ่านชีตไม่ได้ - {e}")
            sheets[prop] = None

    dates = {p: d["date"] for p, d in sheets.items() if d}
    distinct = sorted(set(dates.values()))
    date = want or (distinct[0] if len(distinct) == 1 else None)
    if date is None:
        print("\n⚠️  ชีตแต่ละอันถือคนละวัน - ระบุวันที่เองเป็น argument:")
        for p, d in dates.items():
            print(f"     {SHEETS[p][0]:<12} {d}")
        return

    day = datetime.strptime(date, "%Y-%m-%d")
    print(f"\n{'='*78}")
    print(f"สรุปตามคอลัมน์ — ST Files {day.strftime('%-d %b %Y')}")
    print("="*78)
    if want and want not in distinct:
        # Hard stop, not a warning. These workbooks hold ONE pasted export
        # each, so a day they don't hold cannot be reconstructed from them -
        # and a table comparing our 23rd against the sheets' 25th looks
        # exactly as authoritative as a real one while being nonsense.
        print(f"\n⛔ ชีตไม่ได้เก็บข้อมูลวันที่ {want} — ตอนนี้ถือวันที่ {', '.join(distinct)}")
        print("   ชีตแต่ละอันเก็บได้ครั้งละ 1 วันเท่านั้น (ทับของเดิมทุกครั้งที่วางข้อมูลใหม่)")
        print("   จึงย้อนไปเทียบวันที่ผ่านมาไม่ได้ ต้องรันในวันที่ชีตยังถือข้อมูลวันนั้นอยู่")
        return
    print(f"ชีตทั้ง 8 ถือข้อมูลวันที่ {day.strftime('%-d %b %Y')} ตรงกัน")

    # Our side, via the very same call the ST Files page's own history table
    # uses - so this check can never drift from what the app shows.
    ours = {}
    for prop in SHEETS:
        try:
            rows = await sync_service.get_st_files_list(prop)
            ours[prop] = next((r for r in rows if r.get("date") == date), None)
        except Exception as e:
            print(f"  !! {SHEETS[prop][0]}: อ่าน st_files_sync ไม่ได้ - {e}")
            ours[prop] = None

    stamps = [local(v["synced_at"], p) for p, v in ours.items() if v and v.get("synced_at")]
    if stamps:
        print(f"เทียบ snapshot ของเรา (จับเวลา {min(stamps)} – {max(stamps)}) กับชีต\n")

    # ---- per-column summary (the table asked for) -------------------------
    w = 16
    print(f"{'คอลัมน์':<{w}}{'ตรง':<8}หมายเหตุ")
    print("-"*78)
    detail = {}
    for label, key in METRICS:
        ok, notes = 0, []
        for prop, (short, _) in SHEETS.items():
            sh, ou = sheets.get(prop), ours.get(prop)
            if not sh or ou is None:
                notes.append(f"{short} ไม่มีข้อมูล")
                continue
            sv, ov = as_int(sh["master"].get(label)), ou.get(key)
            if sv == ov:
                ok += 1
            else:
                notes.append(f"{short} {ov - sv:+d}")
                detail.setdefault(prop, []).append((label, ov, sv))
        note = "✅" if ok == len(SHEETS) else ", ".join(notes)
        print(f"{label:<{w}}{f'{ok}/{len(SHEETS)}':<8}{note}")

    total = len(METRICS) * len(SHEETS)
    bad = sum(len(v) for v in detail.values())
    print("-"*78)
    print(f"ตรงกัน {total - bad}/{total} ช่อง" + ("  — ทุกช่องตรงหมด ✅" if not bad else ""))

    if detail:
        print(f"\nรายละเอียดที่ต่าง (เรา / ชีต):")
        for prop, items in detail.items():
            print(f"  {SHEETS[prop][0]:<12} " + ",  ".join(f"{lbl} {o}/{s}" for lbl, o, s in items))

asyncio.run(main())
