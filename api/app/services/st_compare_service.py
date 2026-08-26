"""ST Files verification: our stored numbers vs each property's own
"<Name>-ST" Google Sheet, which is the ground truth for what actually gets
filed.

Deliberately NOT wired into email_templates / Admin > Email Template. This is
a temporary monitoring feed while the new system is being validated, not a
business email anyone should be editing the design of - and a template edit
that dropped the table would silently turn the daily check into a blank page.
Recipient and schedule are constants here; delete this module and its job in
main.py to switch the monitoring off.

Two things about the source sheets drive the whole design:

1. Each workbook holds exactly ONE pasted MEWS export, for whatever day was
   last put into it - it is not a running log. So the date is READ FROM THE
   SHEETS rather than assumed, and a day they no longer hold simply cannot be
   compared.
2. The Master tab mixes TWO different MEWS exports. Spaces/Occupied/House
   uses/Out of order/Availability/Customers/Arrivals/Departures come from the
   Availability report (`Parameters` tab); Complimentary is computed off the
   `Reservation` tab, a separate Reservation report. Only the first group is
   compared cell-for-cell here; Complimentary is compared as a total.
"""
import asyncio
import io
import logging
import re
from collections import OrderedDict
from datetime import datetime, timedelta, timezone

import httpx
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Where the daily monitoring mail goes. One address on purpose - see the
# module docstring for why this is not an Admin-editable recipient list.
MONITOR_RECIPIENT = "khemmarin.k@naraihospitality.com"

# Asia/Bangkok, evaluated the same way every other scheduled job in main.py
# is. 03:00 sits safely after both inputs are ready: the property sheets are
# generated 01:20-02:25, and our own ST Files import runs 00:20-02:03.
SEND_HOUR = 3
SEND_MINUTE = 0

# Same IDs as the links the front-office team maintains.
SHEETS = OrderedDict([
    ("Lub d Bangkok Chinatown",       ("Chinatown", "1npf-d74wYYwsQk9LrNUyYJxnCHH0dsLVuxELJo_CqUM")),
    ("Lub d Bangkok Siam",            ("Siam",      "1JFEQcs1lz62KSIYPuzJhY8o_nP_4PEGdZxmeKrqELVc")),
    ("Lub d Koh Samui Chaweng Beach", ("Samui",     "1DwDsPAajjFH5Fbe_-6Jcl-3ds0q-DCkyGVZr2sRpSw4")),
    ("Lub d Koh Tao Tanote Bay",      ("Koh Tao",   "1-48rztEk1J0TmOiFgPfTE9I_UB0Ba9-RvxOKj8aJyG0")),
    ("Lub d Philippines Makati",      ("Makati",    "1ykTLblIv9dIXEDzdivzPr-z6IWDr_asdeCTn0evut3k")),
    ("Lub d Phuket Patong",           ("Patong",    "1X8mh5Hlvcl6hZhO4Z-eyqdZeGQSFwbYjwHTjzMxcEEs")),
    ("Lub d Siem Reap",               ("Siem Reap", "17siV7sMIT5GsW9x8LdsWBVjh1DcSUTchX-WwiS09FpY")),
    ("Marasca Samui",                 ("Marasca",   "1r2i50lPT8VOFKsjqnSgphawAQvgb3S8yWafyGpVIOCs")),
])

# (Master row label, our st_files_sync key). Siam and Siem Reap label the last
# row "Complimentary Room" where the others say "Complimentary".
METRICS = [
    ("Spaces",        "spaces"),
    ("Occupied",      "occupied"),
    ("House uses",    "house_use"),
    ("Out of order",  "out_of_order"),
    ("Availability",  "availability"),
    ("Customers",     "customers"),
    ("Arrivals",      "arrivals"),
    ("Departures",    "departures"),
    ("Complimentary", "complimentary"),
]

# Manila is UTC+8; every other property is UTC+7.
TZ_OFFSET = {"Lub d Philippines Makati": 8}


def _parse_master(content: bytes) -> dict:
    """Master tab as {label: value}, plus the report date and the timestamp
    MEWS generated the underlying Availability export."""
    wb = load_workbook(io.BytesIO(content), data_only=True)

    master = {}
    ws = wb["Master"]
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()
        if label:
            master[label] = ws.cell(row, 2).value
    if "Complimentary" not in master and "Complimentary Room" in master:
        master["Complimentary"] = master["Complimentary Room"]

    # Prefer the Availability parameters tab - several sheets also carry a
    # "Parameters Reservation"/"Parameters-Reservation" tab for the OTHER
    # export, whose Created time would be the wrong one to report.
    created = None
    params = next((wb[n] for n in ("Parameters", "Parameters-Availability") if n in wb.sheetnames), None)
    if params is None:
        params = next((wb[n] for n in wb.sheetnames
                       if n.startswith("Parameters") and "eservation" not in n), None)
    if params is not None:
        for row in range(1, 21):
            if str(params.cell(row, 1).value or "").strip() == "Created":
                created = params.cell(row, 2).value

    report_date = master.get("Date")
    if isinstance(report_date, datetime):
        report_date = report_date.strftime("%Y-%m-%d")
    return {"master": master, "date": report_date, "created": created}


async def _fetch_sheets() -> dict:
    """All 8 Master tabs, fetched concurrently over the workbooks' public
    export URL. A sheet that fails comes back as None rather than failing the
    whole run - one unreachable workbook should still leave seven comparable."""
    async def one(prop, sheet_id):
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        try:
            async with httpx.AsyncClient(follow_redirects=True, timeout=90) as client:
                r = await client.get(url)
                r.raise_for_status()
            return prop, _parse_master(r.content), None
        except Exception as e:
            logger.warning(f"ST compare: could not read {prop}'s sheet: {e}")
            return prop, None, str(e)

    results = await asyncio.gather(*(one(p, sid) for p, (_, sid) in SHEETS.items()))
    return {p: {"data": d, "error": e} for p, d, e in results}


def _as_int(v):
    """A blank Complimentary cell (Siam, Siem Reap) means zero, not missing."""
    if v is None or v == "":
        return 0
    return int(v) if isinstance(v, (int, float)) else v


def _local(ts: str, prop: str) -> str:
    if not ts:
        return ""
    ts = re.sub(r"\.(\d+)", lambda m: "." + m.group(1)[:6].ljust(6, "0"), ts)
    off = TZ_OFFSET.get(prop, 7)
    return datetime.fromisoformat(ts).astimezone(timezone(timedelta(hours=off))).strftime("%d %b %H:%M")


async def build_comparison(want_date: str = None) -> dict:
    """The whole check, as data. `status` is one of:
      ok             - comparable, see columns/grid
      no_sheet_date  - the sheets disagree with each other about the date
      not_held       - a date was asked for that the sheets no longer hold
    """
    from app.services.sync_service import sync_service

    sheets = await _fetch_sheets()
    dates = {p: v["data"]["date"] for p, v in sheets.items() if v["data"]}
    distinct = sorted({d for d in dates.values() if d})

    if want_date and want_date not in distinct:
        # Hard stop rather than a warning: these workbooks hold one pasted
        # export each, so a day they don't hold cannot be reconstructed - and
        # a table comparing our 23rd against the sheets' 25th reads exactly as
        # authoritative as a real one while being nonsense.
        return {"status": "not_held", "want": want_date, "distinct": distinct, "dates": dates}
    if not want_date and len(distinct) != 1:
        return {"status": "no_sheet_date", "distinct": distinct, "dates": dates}

    date = want_date or distinct[0]

    ours = {}
    for prop in SHEETS:
        try:
            rows = await sync_service.get_st_files_list(prop)
            ours[prop] = next((r for r in rows if r.get("date") == date), None)
        except Exception as e:
            logger.warning(f"ST compare: could not read st_files_sync for {prop}: {e}")
            ours[prop] = None

    columns, grid, detail = [], {}, {}
    for label, key in METRICS:
        ok, notes = 0, []
        for prop, (short, _) in SHEETS.items():
            sh, ou = sheets[prop]["data"], ours.get(prop)
            if not sh or ou is None:
                notes.append(f"{short} ไม่มีข้อมูล")
                grid.setdefault(prop, {})[label] = (None, None)
                continue
            sv, ov = _as_int(sh["master"].get(label)), ou.get(key)
            grid.setdefault(prop, {})[label] = (ov, sv)
            if sv == ov:
                ok += 1
            else:
                notes.append(f"{short} {ov - sv:+d}")
                detail.setdefault(prop, []).append((label, ov, sv))
        columns.append({"label": label, "matched": ok, "total": len(SHEETS), "notes": notes})

    stamps = sorted(s for s in (_local((v or {}).get("synced_at", ""), p) for p, v in ours.items()) if s)
    total = len(METRICS) * len(SHEETS)
    mismatched = sum(len(v) for v in detail.values())
    return {
        "status": "ok",
        "date": date,
        "columns": columns,
        "grid": grid,
        "detail": detail,
        "total_cells": total,
        "matched_cells": total - mismatched,
        "window": (stamps[0], stamps[-1]) if stamps else None,
        "sheet_errors": {p: v["error"] for p, v in sheets.items() if v["error"]},
    }


def _title(result: dict) -> str:
    day = datetime.strptime(result["date"], "%Y-%m-%d")
    return f"ST Files {day.strftime('%-d %b %Y')}"


def render_text(result: dict) -> str:
    """Plain-text form - the CLI output, and the email's text/plain part."""
    if result["status"] == "not_held":
        return (f"⛔ ชีตไม่ได้เก็บข้อมูลวันที่ {result['want']} — ตอนนี้ถือวันที่ "
                f"{', '.join(result['distinct'])}\n"
                "   ชีตแต่ละอันเก็บได้ครั้งละ 1 วันเท่านั้น (ทับของเดิมทุกครั้งที่วางข้อมูลใหม่)\n"
                "   จึงย้อนไปเทียบวันที่ผ่านมาไม่ได้")
    if result["status"] == "no_sheet_date":
        lines = ["⚠️  ชีตแต่ละอันถือคนละวัน - ยังเทียบไม่ได้:"]
        for p, d in result["dates"].items():
            lines.append(f"     {SHEETS[p][0]:<12} {d}")
        return "\n".join(lines)

    out = ["=" * 78, f"สรุปตามคอลัมน์ — {_title(result)}", "=" * 78]
    day = datetime.strptime(result["date"], "%Y-%m-%d")
    out.append(f"ชีตทั้ง {len(SHEETS)} ถือข้อมูลวันที่ {day.strftime('%-d %b %Y')} ตรงกัน")
    if result["window"]:
        out.append(f"เทียบ snapshot ของเรา (จับเวลา {result['window'][0]} – {result['window'][1]}) กับชีต")
    out += ["", f"{'คอลัมน์':<16}{'ตรง':<8}หมายเหตุ", "-" * 78]
    for c in result["columns"]:
        note = "✅" if c["matched"] == c["total"] else ", ".join(c["notes"])
        score = f"{c['matched']}/{c['total']}"
        out.append(f"{c['label']:<16}{score:<8}{note}")
    out.append("-" * 78)
    out.append(f"ตรงกัน {result['matched_cells']}/{result['total_cells']} ช่อง"
               + ("  — ทุกช่องตรงหมด ✅" if result["matched_cells"] == result["total_cells"] else ""))
    if result["detail"]:
        out.append("")
        out.append("รายละเอียดที่ต่าง (เรา / ชีต):")
        for prop, items in result["detail"].items():
            out.append(f"  {SHEETS[prop][0]:<12} " + ",  ".join(f"{l} {o}/{s}" for l, o, s in items))
    for prop, err in (result.get("sheet_errors") or {}).items():
        out.append(f"  !! {SHEETS[prop][0]}: อ่านชีตไม่ได้ - {err}")
    return "\n".join(out)


_TD = "padding:6px 10px;border:1px solid #e2e8f0;font-size:13px;"
_TH = "padding:6px 10px;border:1px solid #e2e8f0;font-size:11px;font-weight:700;background:#f8fafc;text-align:left;"


def render_html(result: dict) -> str:
    """The email body. Inline styles only, plain tables - this has to survive
    Outlook and Gmail, not just a browser."""
    if result["status"] != "ok":
        return f"<pre style=\"font-family:ui-monospace,monospace;font-size:13px\">{render_text(result)}</pre>"

    day = datetime.strptime(result["date"], "%Y-%m-%d")
    perfect = result["matched_cells"] == result["total_cells"]
    banner_bg, banner_fg = ("#dcfce7", "#166534") if perfect else ("#fef9c3", "#854d0e")

    h = [f"""<div style="font-family:-apple-system,Segoe UI,Helvetica,Arial,sans-serif;color:#0f172a;max-width:900px">
<h2 style="margin:0 0 4px">สรุปตามคอลัมน์ — {_title(result)}</h2>
<div style="background:{banner_bg};color:{banner_fg};padding:8px 12px;border-radius:6px;
 font-weight:700;font-size:14px;margin:10px 0">
 ตรงกัน {result['matched_cells']}/{result['total_cells']} ช่อง{' — ทุกช่องตรงหมด' if perfect else ''}</div>
<p style="font-size:13px;color:#475569;margin:6px 0 14px">
ชีตทั้ง {len(SHEETS)} ถือข้อมูลวันที่ {day.strftime('%-d %b %Y')} ตรงกัน"""]
    if result["window"]:
        h.append(f" · เทียบ snapshot ของเรา (จับเวลา {result['window'][0]} – {result['window'][1]}) กับชีต")
    h.append("</p>")

    # --- per-column summary (the table asked for) --------------------------
    h.append(f'<table style="border-collapse:collapse;margin-bottom:22px"><tr>'
             f'<th style="{_TH}">คอลัมน์</th><th style="{_TH}">ตรง</th><th style="{_TH}">หมายเหตุ</th></tr>')
    for c in result["columns"]:
        good = c["matched"] == c["total"]
        note = "✅" if good else ", ".join(c["notes"])
        colour = "" if good else "color:#b45309;font-weight:700;"
        h.append(f'<tr><td style="{_TD}font-weight:600">{c["label"]}</td>'
                 f'<td style="{_TD}{colour}">{c["matched"]}/{c["total"]}</td>'
                 f'<td style="{_TD}{colour}">{note}</td></tr>')
    h.append("</table>")

    # --- full grid ---------------------------------------------------------
    h.append('<h3 style="margin:0 0 8px;font-size:15px">ตารางเต็ม — ระบบเรา / ชีต</h3>')
    h.append(f'<table style="border-collapse:collapse"><tr><th style="{_TH}">Property</th>')
    for label, _ in METRICS:
        h.append(f'<th style="{_TH}">{label}</th>')
    h.append("</tr>")
    for prop, (short, _) in SHEETS.items():
        h.append(f'<tr><td style="{_TD}font-weight:600;white-space:nowrap">{short}</td>')
        for label, _k in METRICS:
            ov, sv = result["grid"].get(prop, {}).get(label, (None, None))
            if ov is None and sv is None:
                cell, style = "—", "color:#94a3b8;"
            elif ov == sv:
                cell, style = f"✓ {ov}", "color:#475569;"
            else:
                cell, style = f"<b>{ov} / {sv}</b>", "background:#fef3c7;color:#92400e;"
            h.append(f'<td style="{_TD}{style}">{cell}</td>')
        h.append("</tr>")
    h.append("</table>")

    h.append('<p style="font-size:11px;color:#94a3b8;margin-top:18px">'
             'ระบบเรา / ชีต — ✓ คือตรงกัน · เมลเฝ้าระวังอัตโนมัติจาก NHGOne '
             '(ไม่ได้ตั้งค่าใน Admin &gt; Email Template)</p></div>')
    return "".join(h)
