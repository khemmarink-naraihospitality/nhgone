"""GL account reconciliation and bank-vs-GL matching. Both operate purely on
uploaded Excel files with no MEWS/Supabase involvement, so they live in their
own module rather than sync_service.py."""
import io
import re
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_MISMATCH_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


def _find_header_row(raw: pd.DataFrame, required_cols: list, scan_rows: int = 30) -> int:
    """GL exports carry a few title/metadata rows above the real column
    headers - scan for the row that actually contains every required
    column name, rather than assuming a fixed row number."""
    for idx in range(min(scan_rows, len(raw))):
        vals = raw.iloc[idx].astype(str).str.strip().tolist()
        if all(col in vals for col in required_cols):
            return idx
    raise ValueError(f"Could not find a header row containing {required_cols} in the first {scan_rows} rows")


def _clean_account_code(v) -> str:
    """Excel silently stores a text code like '30005' as a float (30005.0)
    once read - strip the trailing '.0' rather than losing the string form
    a leading-zero code would need."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _safe_filename(name: str) -> str:
    name = re.sub(r'[\\/*?:"<>|]', "_", str(name)).strip()
    name = re.sub(r"\s+", " ", name)
    return name[:180]


def _exact_offsets(group: pd.DataFrame, amount_col: str):
    """Greedy pairwise matching within one account: pairs up any two
    transactions whose amounts sum to zero (a charge and its exact reversal)
    into 'Offset'; everything left over is 'Outstanding'. First-match
    greedy - a specific reversal can't always be told apart from another of
    the same amount, so this doesn't try to guess intent beyond "some
    pairing that zeroes out"."""
    g = group.reset_index(drop=True)
    used = set()
    matched_rows = []
    for i in range(len(g)):
        if i in used:
            continue
        a = g.at[i, amount_col]
        if pd.isna(a):
            continue
        for j in range(i + 1, len(g)):
            if j in used:
                continue
            b = g.at[j, amount_col]
            if pd.isna(b):
                continue
            if round(float(a) + float(b), 2) == 0:
                matched_rows.extend([g.iloc[i], g.iloc[j]])
                used.update([i, j])
                break
    offset_df = pd.DataFrame(matched_rows, columns=g.columns) if matched_rows else pd.DataFrame(columns=g.columns)
    outstanding_df = g.drop(index=list(used)).reset_index(drop=True)
    return offset_df.reset_index(drop=True), outstanding_df


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)
    for col_num, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(1, col_num, col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for row_num, row in enumerate(df.itertuples(index=False), start=2):
        for col_num, value in enumerate(row, start=1):
            ws.cell(row_num, col_num, None if pd.isna(value) else value)
    ws.freeze_panes = "A2"
    for col_num, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_num)
        if col_name == "Account Code":
            ws.column_dimensions[letter].width = 14
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_num).number_format = "@"
        elif col_name in ("Base Amount", "Bank amount"):
            ws.column_dimensions[letter].width = 16
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_num).number_format = "#,##0.00"
        elif "Date" in str(col_name):
            ws.column_dimensions[letter].width = 14
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col_num).number_format = "yyyy-mm-dd"
        else:
            ws.column_dimensions[letter].width = min(max(len(str(col_name)) + 2, 10), 35)
    return ws


def run_gl_split(file_bytes: bytes) -> bytes:
    """Splits a GL Account Detail export into one workbook per Account Code
    (All Transactions / Outstanding / Offset sheets) plus a summary
    workbook with a per-account balance check, all zipped together.
    Returns the zip file's raw bytes."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    sheet_name = "GL" if "GL" in xls.sheet_names else xls.sheet_names[0]
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    header_row = _find_header_row(raw, ["Account Code", "Base Amount"])

    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
    df.columns = df.columns.astype(str).str.strip()
    df = df.loc[:, ~df.columns.duplicated()].copy()
    if "Account Description" not in df.columns and "Description" in df.columns:
        df = df.rename(columns={"Description": "Account Description"})

    required = ["Account Code", "Account Description", "Base Amount"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    df["Account Code Clean"] = df["Account Code"].apply(_clean_account_code)
    df = df[(df["Account Code Clean"] != "") & (df["Account Code Clean"].str.lower() != "nan")].copy()
    df["Base Amount"] = pd.to_numeric(df["Base Amount"], errors="coerce")

    groups = list(df.groupby("Account Code Clean", sort=True))
    if not groups:
        raise ValueError("No rows with a valid Account Code were found")

    zip_buf = io.BytesIO()
    summary_rows = []
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for code, group in groups:
            group = group.drop(columns=["Account Code Clean"]).reset_index(drop=True)
            group["Account Code"] = code
            desc_values = group["Account Description"].dropna().astype(str).str.strip()
            desc_values = desc_values[desc_values != ""]
            desc = desc_values.iloc[0] if len(desc_values) else "No Description"

            all_tx = group.copy()
            offset_df, outstanding_df = _exact_offsets(all_tx, "Base Amount")

            wb = Workbook()
            wb.remove(wb.active)
            _write_sheet(wb, "All Transactions", all_tx)
            _write_sheet(wb, "Outstanding", outstanding_df)
            _write_sheet(wb, "Offset", offset_df)
            file_buf = io.BytesIO()
            wb.save(file_buf)
            z.writestr(_safe_filename(f"{code} {desc}") + ".xlsx", file_buf.getvalue())

            summary_rows.append({
                "Account Code": code,
                "Account Description": desc,
                "All Transactions Rows": len(all_tx),
                "Outstanding Rows": len(outstanding_df),
                "Offset Rows": len(offset_df),
                "Balance Check": round(
                    all_tx["Base Amount"].sum(skipna=True)
                    - outstanding_df["Base Amount"].sum(skipna=True)
                    - offset_df["Base Amount"].sum(skipna=True), 2),
            })

        summary_df = pd.DataFrame(summary_rows)
        wb = Workbook()
        wb.remove(wb.active)
        _write_sheet(wb, "Summary", summary_df)
        summary_buf = io.BytesIO()
        wb.save(summary_buf)
        z.writestr("_Summary_Check.xlsx", summary_buf.getvalue())

    return zip_buf.getvalue()


def run_bank_gl_match(bank_bytes: bytes, gl_bytes: bytes) -> bytes:
    """Matches Bank Statement rows to GL Statement rows by amount (rounded
    to 2dp) and a same-day-or-1-day-off date tolerance. Matching is greedy
    and index-based - never by value - so two different transactions that
    happen to share the same date and amount are still tracked and paired
    off independently rather than collapsed into one. Returns one workbook's
    raw bytes: Summary, each side's Outstanding (unmatched) items with
    cross-matching amounts highlighted for review, the full input, and the
    matched pairs."""
    bank_df = pd.read_excel(io.BytesIO(bank_bytes))
    gl_df = pd.read_excel(io.BytesIO(gl_bytes))
    bank_df.columns = [str(c).strip().replace("\n", " ") for c in bank_df.columns]
    gl_df.columns = [str(c).strip().replace("\n", " ") for c in gl_df.columns]

    for col, df, label in [("TranDate", bank_df, "Bank Statement"), ("Bank amount", bank_df, "Bank Statement"),
                            ("Transaction Date", gl_df, "GL Statement"), ("Base Amount", gl_df, "GL Statement")]:
        if col not in df.columns:
            raise ValueError(f"{label} is missing required column '{col}'")

    bank_df["TranDate"] = pd.to_datetime(bank_df["TranDate"], errors="coerce", dayfirst=True).dt.date
    bank_df["Bank amount"] = pd.to_numeric(bank_df["Bank amount"], errors="coerce")
    gl_df["Transaction Date"] = pd.to_datetime(gl_df["Transaction Date"], errors="coerce", dayfirst=True).dt.date
    gl_df["Base Amount"] = pd.to_numeric(gl_df["Base Amount"], errors="coerce")
    bank_df = bank_df.dropna(subset=["TranDate", "Bank amount"]).reset_index(drop=True)
    gl_df = gl_df.dropna(subset=["Transaction Date", "Base Amount"]).reset_index(drop=True)

    used_gl_idx, used_bank_idx = set(), set()
    bank_match_rows, gl_match_rows = [], []
    for bi, b in bank_df.iterrows():
        amt = round(float(b["Bank amount"]), 2)
        dt = b["TranDate"]
        for gi, g in gl_df.iterrows():
            if gi in used_gl_idx:
                continue
            if pd.isna(g["Base Amount"]) or round(float(g["Base Amount"]), 2) != amt:
                continue
            if abs((g["Transaction Date"] - dt).days) > 1:
                continue
            used_gl_idx.add(gi)
            used_bank_idx.add(bi)
            bank_match_rows.append(b)
            gl_match_rows.append(g)
            break

    bank_match_df = pd.DataFrame(bank_match_rows, columns=bank_df.columns) if bank_match_rows else pd.DataFrame(columns=bank_df.columns)
    gl_match_df = pd.DataFrame(gl_match_rows, columns=gl_df.columns) if gl_match_rows else pd.DataFrame(columns=gl_df.columns)
    bank_out = bank_df.drop(index=list(used_bank_idx)).reset_index(drop=True)
    gl_out = gl_df.drop(index=list(used_gl_idx)).reset_index(drop=True)

    same_amounts = set(bank_out["Bank amount"].round(2)) & set(gl_out["Base Amount"].round(2))

    summary_df = pd.DataFrame({
        "Type": ["Bank Outstanding", "GL Outstanding", "Bank_Full", "GL_Full", "Bank Match", "GL Match"],
        "Total Transactions": [len(bank_out), len(gl_out), len(bank_df), len(gl_df), len(bank_match_df), len(gl_match_df)],
        "Total Amount": [
            round(float(bank_out["Bank amount"].sum()), 2), round(float(gl_out["Base Amount"].sum()), 2),
            round(float(bank_df["Bank amount"].sum()), 2), round(float(gl_df["Base Amount"].sum()), 2),
            round(float(bank_match_df["Bank amount"].sum()), 2) if len(bank_match_df) else 0.0,
            round(float(gl_match_df["Base Amount"].sum()), 2) if len(gl_match_df) else 0.0,
        ],
    })

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Summary", summary_df)
    ws_bank_out = _write_sheet(wb, "Bank Outstanding", bank_out)
    ws_gl_out = _write_sheet(wb, "GL Outstanding", gl_out)
    _write_sheet(wb, "Bank_Full", bank_df)
    _write_sheet(wb, "GL_Full", gl_df)
    _write_sheet(wb, "Bank Match", bank_match_df)
    _write_sheet(wb, "GL Match", gl_match_df)

    for ws, df, col_name in [(ws_bank_out, bank_out, "Bank amount"), (ws_gl_out, gl_out, "Base Amount")]:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, col_idx).value
            if val is not None and round(float(val), 2) in same_amounts:
                ws.cell(r, col_idx).fill = _MISMATCH_FILL

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()
