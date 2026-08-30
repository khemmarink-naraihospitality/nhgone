"""RR4 / TM30 verification: our stored register vs each property's own
"RR4-TM30-<Name>-Gen" Google Sheet, which is the ground truth for what
actually gets filed with the authorities.

The RR4/TM30 counterpart to st_compare_service, and deliberately the same
shape - but three things differ, each forced by the source sheets:

1. **Thailand only.** Lub d Siem Reap and Lub d Philippines Makati don't file
   under the Thai Hotel Act and have no generator sheet at all (same exclusion
   sync_service._RR4_TM30_EMAIL_EXCLUDED_PROPERTIES already applies to the
   daily RR4/TM30 digest).

2. **Each property is compared at ITS OWN date**, not one date shared by all -
   unlike the ST sheets, which are all pasted within an hour of each other.
   These windows run to the property's own cutoff hour, and Chinatown's is
   12:15 where everyone else's is ~02:00, so at any given moment Chinatown's
   sheet is a full day behind the rest. Demanding one common date would mean
   never sending a mail at all.

3. **Rows are paired by identity, then compared column by column.** A
   key-based diff that stops at "we found a row with this passport" hides real
   defects (on 2026-08-22 a 4-field key called Patong's TM30 a perfect match
   while a full-field diff surfaced 6 differing rows). So the passport/PID is
   used ONLY to pair the two sides up; every column of a paired row is then
   compared, and a name that changed shows up as a column difference rather
   than as two unmatched rows.

Known drift is reported separately from real differences - see _KNOWN_DRIFT
below. Those four patterns have each been investigated and confirmed to be
the sheet or MEWS moving on after the fact, not our bug; counting them as
mismatches every single day would bury the differences that do matter.
"""
import asyncio
import io
import logging
import re
from collections import OrderedDict
from datetime import datetime, timedelta, timezone

import httpx
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# The six Thai properties' generator workbooks. Marasca is the "MRCS"-style
# sheet, NOT `1tY_kX...`/MRCKY-Gen, whose ImportInhouse was empty when checked.
SHEETS = OrderedDict([
    ("Lub d Bangkok Chinatown",       ("Chinatown", "1qT4ZClqvTLVUW9Bc4Oaxx2oy6u8QZo0dCVlyM35JRy4")),
    ("Lub d Bangkok Siam",            ("Siam",      "1liiB8tqYGCgAyKqDsRCCaZSonubHgnMr2YSZQr7-SlQ")),
    ("Lub d Koh Samui Chaweng Beach", ("Samui",     "1nanCOqRnRjiFzkQ_l0RyZqJ0oZ_LGJ6-RTB0qDMTAyg")),
    ("Lub d Koh Tao Tanote Bay",      ("Koh Tao",   "1akGkOIoHKURs6DihwkCRw5zx37HkWVlaYFSdjj-KI6c")),
    ("Lub d Phuket Patong",           ("Patong",    "1XKfU7pSyMwSFIiq7g1wKlKJB_gW9d1JtTuniahqBja8")),
    ("Marasca Samui",                 ("Marasca",   "1YZD0CYpaOwxSiHLa7iH_7bK3ED5CAIhdR2GizwKKwuI")),
])

# rowNo is excluded from the comparison: both sides renumber their own rows
# from 1, and the two exports don't emit guests in the same order (the sheet
# lists a room's unnamed occupant slot first, we don't), so it would report a
# difference on almost every row while meaning nothing.
_SKIP_RR4_COLUMNS = {"row_no"}

# Differences that have each been chased down and confirmed as the world
# moving on after the sheet was generated, not a defect on our side. Reported
# in their own column so the "real differences" number stays meaningful.
_KNOWN_DRIFT = {
    "time_check_in":
        "MEWS wrote ActualStartUtc at :59 seconds, right after the sheet was generated (sheet is exactly 1 minute behind)",
    "date_check_out":
        "Guest checked out earlier than scheduled, after the sheet was generated (ours is ahead of the sheet)",
    "check_out_date":
        "Guest checked out earlier than scheduled, after the sheet was generated (ours is ahead of the sheet)",
    "birth_date":
        "Sheet prints 30/12/1899 when MEWS has no birth date (Excel's render of an empty value) - ours leaves it blank, which is correct",
}


def _norm(v) -> str:
    """One cell, normalized for comparison. Strips the leading apostrophe the
    sheets use to force Text formatting (our export carries it too, on the
    date columns only - see the RR4 importer's own rule), collapses 63 vs
    63.0 vs "63", and treats None and "" as the same empty."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s.startswith("'"):
        s = s[1:].strip()
    return s


def _dmy(s: str):
    """(y, m, d) from a dd/mm/yyyy string - Buddhist or Christian era, since
    both sides of any one column always use the same one."""
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    return (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else None


def _hhmm(s: str):
    """Minutes-since-midnight from the sheets' "HH.MM" check-in time."""
    m = re.match(r"^(\d{1,2})[.:](\d{2})$", s)
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def _is_known_drift(key: str, ours: str, sheet: str) -> bool:
    if key == "time_check_in":
        a, b = _hhmm(ours), _hhmm(sheet)
        return a is not None and b is not None and b - a == 1
    if key in ("date_check_out", "check_out_date"):
        a, b = _dmy(ours), _dmy(sheet)
        return a is not None and b is not None and a < b
    if key == "birth_date":
        return ours == "" and sheet == "30/12/1899"
    return False


def _pair_key(row: dict, kind: str) -> tuple:
    """What identifies the same guest on both sides. Passport first (the one
    field the filing itself is keyed on), then Thai national ID, then the
    name - and for MEWS's unnamed occupant slots, which carry none of the
    three, the room and check-in they were booked under."""
    if kind == "rr4":
        pp, pid = _norm(row.get("passport")).upper(), _norm(row.get("pid")).upper()
        if pp:
            return ("P", pp)
        if pid:
            return ("I", pid)
        name = (_norm(row.get("name_en")) + "|" + _norm(row.get("surname_en"))).upper()
        if name != "|":
            return ("N", name)
        # Deliberately NOT keyed on time_check_in: the two exports routinely
        # disagree by a minute on it (_is_known_drift has a rule for exactly
        # that), and a field known to drift cannot also be an identity. When
        # it was part of this key the drift stopped the row pairing at all,
        # so one unnamed slot surfaced as "only ours" AND "only sheet"
        # instead of as the known drift it is - Patong room 2313 on
        # 2026-08-26 (ours 06.42, sheet 06.43) was doing exactly that. Room +
        # check-in date still separates the slots; two of them sharing even
        # that are paired by content in _index.
        return ("X", _norm(row.get("room_no")), _norm(row.get("date_check_in")))
    pp = _norm(row.get("passport_no")).upper()
    if pp:
        return ("P", pp)
    return ("N", (_norm(row.get("first_name")) + "|" + _norm(row.get("last_name"))).upper())


def _index(rows: list, kind: str, columns: list) -> dict:
    """Rows by pair key, with an occurrence counter appended so two guests
    sharing a passport (or two unnamed slots in the same room) stay distinct
    instead of one silently overwriting the other.

    Where a key DOES repeat, the duplicates are ordered by their own compared
    values rather than by the order the export happened to emit them. The two
    sides genuinely do emit rows in different orders (see the module
    docstring), so numbering by first-seen paired a guest's first row against
    the sheet's second and reported every column that differs between the
    guest's own two rows as a difference on both of them - twice over, once
    in each direction. Verified 2026-08-26: Chinatown's Andrea Solves Vidal
    (530 @ 21.43 + 404 @ 21.44) and Siam's PATTRAPORN CHANIM (105 @ 13.59 +
    101 @ 14.00) each held identical data on both sides in opposite order,
    and accounted for 4 of the 5 "real" RR4 differences that day. Sorting on
    the compared columns is deterministic and identical on both sides, so
    matching rows line up and genuinely different ones still report.
    """
    groups = {}
    for row in rows:
        groups.setdefault(_pair_key(row, kind), []).append(row)
    out = {}
    for k, group in groups.items():
        if len(group) > 1:
            group = sorted(group, key=lambda r: tuple(_norm(r.get(c)) for c in columns))
        for i, row in enumerate(group, 1):
            out[k + (i,)] = row
    return out


def _compare_rows(ours: list, sheet: list, kind: str, columns: list) -> dict:
    """One property, one register. Pairs the two sides up by identity, then
    compares every column of every paired row."""
    ours_ix, sheet_ix = _index(ours, kind, columns), _index(sheet, kind, columns)
    paired = ours_ix.keys() & sheet_ix.keys()

    diff_rows, drift_rows = 0, 0
    cols, drift_cols, samples = {}, {}, []
    for k in paired:
        o, s = ours_ix[k], sheet_ix[k]
        real, drift = [], []
        for key in columns:
            ov, sv = _norm(o.get(key)), _norm(s.get(key))
            if ov == sv:
                continue
            (drift if _is_known_drift(key, ov, sv) else real).append((key, ov, sv))
        for key, _o, _s in drift:
            drift_cols[key] = drift_cols.get(key, 0) + 1
        for key, _o, _s in real:
            cols[key] = cols.get(key, 0) + 1
        if real:
            diff_rows += 1
            if len(samples) < 3:
                label = (_norm(o.get("name_en")) or _norm(o.get("first_name")) or "—") + " " + \
                        (_norm(o.get("surname_en")) or _norm(o.get("last_name")) or "")
                samples.append((label.strip(), real[:4]))
        elif drift:
            drift_rows += 1

    return {
        "ours": len(ours),
        "sheet": len(sheet),
        "paired": len(paired),
        "only_ours": len(ours_ix.keys() - sheet_ix.keys()),
        "only_sheet": len(sheet_ix.keys() - ours_ix.keys()),
        "clean_rows": len(paired) - diff_rows - drift_rows,
        "diff_rows": diff_rows,
        "drift_rows": drift_rows,
        "cols": cols,
        "drift_cols": drift_cols,
        "samples": samples,
    }


def _parse_sheet(content: bytes) -> dict:
    """One workbook's RR4 and TM30 tabs, plus the window each was exported
    over. The Master tab is the only place every sheet agrees on: A2 is the
    ImportInhouse (RR4) start and B2 the ImportCP (TM30) start. Chinatown also
    carries Parameter-* tabs, but the other five don't."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    ms = wb["Master"]
    rr4_start, tm30_start = ms.cell(2, 1).value, ms.cell(2, 2).value
    date = rr4_start.strftime("%Y-%m-%d") if isinstance(rr4_start, datetime) else None

    # RR4 tab: row 3 is the Thai header, row 4 the English field keys, data
    # from row 5. Columns are located BY that field-key row rather than by
    # position, so a sheet that gains a column doesn't silently shift the
    # whole comparison one to the left.
    from app.services.sync_service import sync_service
    rr4_rows = []
    if "RR4" in wb.sheetnames:
        ws = wb["RR4"]
        grid = [[c.value for c in row] for row in ws.iter_rows(min_row=4, max_col=27)]
        keys = {_norm(v): i for i, v in enumerate(grid[0])} if grid else {}
        by_key = {field: keys.get(field) for _k, _l, field in sync_service._RR4_COLUMNS}
        for row in grid[1:]:
            if _norm(row[0]) == "":
                continue
            rr4_rows.append({
                key: (row[by_key[field]] if by_key.get(field) is not None else None)
                for key, _label, field in sync_service._RR4_COLUMNS
            })

    # TM30 tab: one header row of the government form's own bilingual labels,
    # then data. Nine fixed columns in the export's own order - matched
    # positionally because those labels carry embedded newlines and stray
    # spaces that differ between sheets.
    tm30_rows = []
    if "TM30" in wb.sheetnames:
        ws = wb["TM30"]
        for row in ws.iter_rows(min_row=2, max_col=len(sync_service._TM30_COLUMNS), values_only=True):
            if _norm(row[0]) == "":
                continue
            tm30_rows.append({key: row[i] for i, (key, _label) in enumerate(sync_service._TM30_COLUMNS)})

    wb.close()
    return {
        "date": date,
        "rr4_rows": rr4_rows,
        "tm30_rows": tm30_rows,
        "rr4_window": rr4_start.strftime("%H:%M") if isinstance(rr4_start, datetime) else None,
        "tm30_window": tm30_start.strftime("%H:%M") if isinstance(tm30_start, datetime) else None,
    }


async def _fetch_sheets() -> dict:
    """All six workbooks over their public export URL, concurrently. One that
    fails comes back as an error on that property's row rather than failing
    the run - five comparable properties still beat none."""
    async def one(prop, sheet_id):
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        try:
            async with httpx.AsyncClient(follow_redirects=True, timeout=120) as client:
                r = await client.get(url)
                r.raise_for_status()
            return prop, _parse_sheet(r.content), None
        except Exception as e:
            logger.warning(f"RR4 compare: could not read {prop}'s sheet: {e}")
            return prop, None, str(e)

    results = await asyncio.gather(*(one(p, sid) for p, (_, sid) in SHEETS.items()))
    return {p: {"data": d, "error": e} for p, d, e in results}


def _windows() -> dict:
    """Each property's configured RR4 day window, to be shown next to the one
    the sheet was actually exported over. A stale value here silently
    undercounts the register - a leftover 14:00/12:00 on Chinatown once
    produced 160 rows where the real answer was 241 - and the sheets change
    their window without warning, so this is worth a daily look."""
    from app.services.sync_service import sync_service
    try:
        res = sync_service.supabase.table("property_api_settings").select(
            "property_name, rr4_tm30_day_start_hour, rr4_tm30_day_start_minute").execute()
        return {r["property_name"]: f"{r.get('rr4_tm30_day_start_hour') or 0:02d}:"
                                    f"{r.get('rr4_tm30_day_start_minute') or 0:02d}"
                for r in (res.data or [])}
    except Exception as e:
        logger.warning(f"RR4 compare: could not read property windows: {e}")
        return {}


async def _tm30_windows() -> dict:
    """TM30's own configured window per property - a separate setting from
    RR4's, so the mail has to ask for it separately too. Read through
    _resolve_tm30_day_start rather than the table directly, so this column
    can never disagree with the one the register was actually built on."""
    from app.services.sync_service import sync_service
    out = {}
    for prop in SHEETS:
        try:
            h, m = await sync_service._resolve_tm30_day_start(prop)
            out[prop] = f"{h:02d}:{m:02d}"
        except Exception as e:
            logger.warning(f"RR4 compare: could not read {prop}'s TM30 window: {e}")
    return out


def _local(ts: str) -> str:
    """A stored UTC timestamp as Bangkok wall-clock. Every property here is in
    Thailand, so there is only the one offset to apply."""
    if not ts:
        return ""
    ts = re.sub(r"\.(\d+)", lambda m: "." + m.group(1)[:6].ljust(6, "0"), ts)
    return datetime.fromisoformat(ts).astimezone(timezone(timedelta(hours=7))).strftime("%d %b %H:%M")


async def build_comparison(want_date: str = None) -> dict:
    """The whole check, as data.

    `want_date` (the CLI's optional argument) pins every property to one date
    instead of each following its own sheet - useful for reproducing a past
    run, and it simply reports "the sheet holds a different date" for any property whose sheet
    has since moved on, because these workbooks hold one pasted export each
    and a day they no longer hold cannot be reconstructed.
    """
    from app.routers.rr4 import read_managed_day

    sheets = await _fetch_sheets()
    windows = _windows()
    tm30_windows = await _tm30_windows()
    props = []

    for prop, (short, _sid) in SHEETS.items():
        row = {"property": prop, "short": short, "date": None, "status": "error",
               "note": "", "rr4": None, "tm30": None, "synced_at": "",
               "sheet_rr4_window": "", "sheet_tm30_window": "",
               "our_window": windows.get(prop, ""),
               "our_tm30_window": tm30_windows.get(prop, "")}
        sh = sheets[prop]["data"]
        if not sh:
            row["note"] = f"Could not read sheet — {sheets[prop]['error']}"
            props.append(row)
            continue

        row["date"] = sh["date"]
        row["sheet_rr4_window"] = sh["rr4_window"] or ""
        row["sheet_tm30_window"] = sh["tm30_window"] or ""
        if not sh["date"]:
            row["note"] = "Sheet has no date in Master"
            props.append(row)
            continue
        if want_date and want_date != sh["date"]:
            row["status"] = "other_date"
            row["note"] = f"Sheet holds {sh['date']}, not {want_date}"
            props.append(row)
            continue

        try:
            payload = read_managed_day(prop, sh["date"])
        except Exception as e:
            row["note"] = f"Could not read rr4_tm30_sync — {e}"
            props.append(row)
            continue
        if not payload:
            row["status"] = "missing"
            row["note"] = f"{sh['date']} has not been imported yet"
            props.append(row)
            continue

        from app.services.sync_service import sync_service
        rr4_cols = [k for k, _l, _f in sync_service._RR4_COLUMNS if k not in _SKIP_RR4_COLUMNS]
        tm30_cols = [k for k, _l in sync_service._TM30_COLUMNS]
        row["status"] = "ok"
        row["synced_at"] = _local(payload.get("_synced_at") or "")
        row["rr4"] = _compare_rows((payload.get("rr4") or {}).get("rows", []),
                                   sh["rr4_rows"], "rr4", rr4_cols)
        row["tm30"] = _compare_rows((payload.get("tm30") or {}).get("rows", []),
                                    sh["tm30_rows"], "tm30", tm30_cols)
        props.append(row)

    compared = [p for p in props if p["status"] == "ok"]
    if not compared:
        return {"status": "no_data", "properties": props, "date": want_date}

    # The headline date is whichever day most properties are sitting on -
    # Chinatown's later cutoff routinely leaves it a day behind the other
    # five, and naming its date in the subject line would misdescribe the mail.
    counts = {}
    for p in compared:
        counts[p["date"]] = counts.get(p["date"], 0) + 1
    date = max(counts, key=lambda d: (counts[d], d))

    totals = {}
    for kind in ("rr4", "tm30"):
        totals[kind] = {
            f: sum(p[kind][f] for p in compared)
            for f in ("ours", "sheet", "paired", "only_ours", "only_sheet",
                      "clean_rows", "diff_rows", "drift_rows")
        }
    return {
        "status": "ok",
        "date": date,
        "mixed_dates": len(counts) > 1,
        "properties": props,
        "compared": len(compared),
        "totals": totals,
    }


def _title(result: dict) -> str:
    day = datetime.strptime(result["date"], "%Y-%m-%d")
    # "%-d" (unpadded day) is a glibc extension - it raises ValueError on
    # Windows, which made this whole mail impossible to preview from a dev
    # machine. Formatting the day separately is portable and identical.
    return f"RR4/TM30 {day.day} {day.strftime('%b %Y')}"


def _cols_note(block: dict) -> str:
    if not block["cols"]:
        return "✅"
    return ", ".join(f"{k} ×{n}" for k, n in sorted(block["cols"].items(), key=lambda kv: -kv[1]))


def render_text(result: dict) -> str:
    """Plain-text form - the CLI output, and the email's text/plain part."""
    if result["status"] != "ok":
        out = ["⛔ Not comparable yet, for any property:"]
        for p in result["properties"]:
            out.append(f"     {p['short']:<12} {p['note']}")
        return "\n".join(out)

    out = ["=" * 82, f"Comparison Summary — {_title(result)}", "=" * 82]
    if result["mixed_dates"]:
        out.append("(Each property is compared at its own sheet's date — Chinatown cuts its day at 12:15, so it runs a day behind the rest)")
    out.append("")
    out.append(f"{'Property':<12}{'Date':<12}{'RR4 Ours/Sheet':<16}{'Match':<8}{'Diff':<7}{'TM30 Ours/Sheet':<16}{'Match':<8}Diff")
    out.append("-" * 82)
    for p in result["properties"]:
        if p["status"] != "ok":
            out.append(f"{p['short']:<12}{(p['date'] or '—'):<12}{p['note']}")
            continue
        r, t = p["rr4"], p["tm30"]
        r_count = "{}/{}".format(r["ours"], r["sheet"])
        t_count = "{}/{}".format(t["ours"], t["sheet"])
        out.append(
            f"{p['short']:<12}{p['date']:<12}"
            f"{r_count:<16}{r['clean_rows']:<8}{r['diff_rows']:<7}"
            f"{t_count:<16}{t['clean_rows']:<8}{t['diff_rows']}")
    out.append("-" * 82)

    tr, tt = result["totals"]["rr4"], result["totals"]["tm30"]
    out.append(f"RR4  total {tr['ours']}/{tr['sheet']} rows · paired {tr['paired']} · fully matched "
               f"{tr['clean_rows']} · real diff {tr['diff_rows']} · known drift {tr['drift_rows']} · "
               f"ours only {tr['only_ours']} · sheet only {tr['only_sheet']}")
    out.append(f"TM30 total {tt['ours']}/{tt['sheet']} rows · paired {tt['paired']} · fully matched "
               f"{tt['clean_rows']} · real diff {tt['diff_rows']} · known drift {tt['drift_rows']} · "
               f"ours only {tt['only_ours']} · sheet only {tt['only_sheet']}")

    out += ["", "Columns with real differences (ours / sheet):"]
    any_col = False
    for p in result["properties"]:
        if p["status"] != "ok":
            continue
        for kind, label in (("rr4", "RR4"), ("tm30", "TM30")):
            if p[kind]["cols"]:
                any_col = True
                out.append(f"  {p['short']:<12} {label:<5} {_cols_note(p[kind])}")
    if not any_col:
        out.append("  ✅ None")

    samples = [(p["short"], label, who, diffs)
               for p in result["properties"] if p["status"] == "ok"
               for kind, label in (("rr4", "RR4"), ("tm30", "TM30"))
               for who, diffs in p[kind]["samples"]]
    if samples:
        out += ["", "Example differing rows (ours / sheet):"]
        for short, label, who, diffs in samples:
            out.append(f"  {short:<12} {label:<5} {who}")
            for key, ov, sv in diffs:
                out.append(f"       {key:<22} {ov or '—'}  /  {sv or '—'}")

    out += ["", "Export window (sheet) vs configured window (ours):"]
    for p in result["properties"]:
        bad = (p["sheet_rr4_window"] != p["our_window"]
               or p["sheet_tm30_window"] != p["our_tm30_window"])
        flag = "   ⚠️ mismatch" if bad else ""
        out.append(f"  {p['short']:<12} RR4 sheet {p['sheet_rr4_window'] or '—':<7} ours "
                   f"{p['our_window'] or '—':<7} · TM30 sheet {p['sheet_tm30_window'] or '—':<7} ours "
                   f"{p['our_tm30_window'] or '—':<7}{flag}")
    return "\n".join(out)


def _esc(s) -> str:
    """Guest names, passport numbers and addresses come from MEWS, i.e. from
    whatever a guest typed at check-in - they are never trusted as markup."""
    return (str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


_TD = "padding:6px 10px;border:1px solid #e2e8f0;font-size:13px;"
_TH = "padding:6px 10px;border:1px solid #e2e8f0;font-size:11px;font-weight:700;background:#f8fafc;text-align:left;"
_MUTED = "color:#94a3b8;"
_BAD = "background:#fef3c7;color:#92400e;font-weight:700;"


def _summary_cell(block: dict, window: str = "") -> str:
    """One register's cell in the simplified summary grid - same visual
    pattern as st_compare_service's GridTable: a plain "✓ N" when the two
    sides genuinely agree, a bold highlighted "N / M" with a short note
    otherwise. Known drift doesn't count as a real difference here (that's
    the whole point of tracking it separately), so a day with only drift
    still reads as clean at a glance - the detail tables below still show it.

    `window` is the property's configured TM30 start. A non-midnight one
    files a shorter day than the sheet holds on purpose (Chinatown's 12:15
    drops every guest arriving before noon - 2 to 20 of them on each of the
    seven days measured to 29-Aug-2026), so its shortfall is annotated as
    the configured consequence it is. Deliberately still highlighted rather
    than hidden: unlike _KNOWN_DRIFT, this cannot be verified row by row -
    TM30 carries no check-in column to test each missing guest against - so
    a genuine new miss would look identical, and the number stays in view.
    """
    real = block["diff_rows"] + block["only_ours"] + block["only_sheet"]
    shifted = bool(window) and window != "00:00"
    if real == 0:
        note = f" ({block['drift_rows']} known drift)" if block["drift_rows"] else ""
        return f'<td style="{_TD}color:#475569;">✓ {block["ours"]}{note}</td>'
    bits = []
    if block["diff_rows"]:
        bits.append(f"{block['diff_rows']} differ")
    if block["only_ours"]:
        bits.append(f"{block['only_ours']} ours only")
    if block["only_sheet"]:
        bits.append(f"{block['only_sheet']} sheet only"
                    + (f", expected from the {window} window" if shifted else ""))
    return f'<td style="{_TD}{_BAD}"><b>{block["ours"]} / {block["sheet"]}</b> ({", ".join(bits)})</td>'


def render_summary_table(result: dict) -> str:
    """Per-property counts - the <<SummaryTable>> token. Deliberately just
    four columns (Property/Date/RR4/TM30), matching st_compare_service's
    plain grid rather than the earlier 10-column version - the column-level
    and row-level breakdowns still exist, just moved into the two detail
    tables below where someone actually chasing a difference will look."""
    if result["status"] != "ok":
        return f'<pre style="font-family:ui-monospace,monospace;font-size:12px">{render_text(result)}</pre>'

    h = [f'<table style="border-collapse:collapse;width:100%"><tr>'
         f'<th style="{_TH}">Property</th><th style="{_TH}">Date</th>'
         f'<th style="{_TH}">RR4</th><th style="{_TH}">TM30</th></tr>']
    for p in result["properties"]:
        h.append(f'<tr><td style="{_TD}font-weight:600;white-space:nowrap">{p["short"]}</td>')
        if p["status"] != "ok":
            h.append(f'<td style="{_TD}" colspan="3">'
                     f'<span style="{_BAD}padding:2px 6px;border-radius:4px">{p["note"]}</span></td></tr>')
            continue
        h.append(f'<td style="{_TD}{_MUTED}white-space:nowrap">{p["date"]}</td>')
        h.append(_summary_cell(p["rr4"]))
        h.append(_summary_cell(p["tm30"], p.get("our_tm30_window", "")))
        h.append("</tr>")

    tr, tt = result["totals"]["rr4"], result["totals"]["tm30"]
    h.append(f'<tr style="background:#f1f5f9"><td style="{_TD}font-weight:700">Total</td>'
             f'<td style="{_TD}{_MUTED}">{result["compared"]} properties</td>')
    h.append(_summary_cell(tr))
    h.append(_summary_cell(tt))
    h.append("</tr></table>")
    h.append(f'<p style="font-size:11px;color:#94a3b8;margin:6px 0 0">'
             f'Rows are paired by passport/ID number, then every column of each paired row is compared '
             f'(except the row number, which both sides renumber on their own). '
             f'Rows with no name (an MEWS-booked slot not yet linked to a guest profile) are counted here '
             f'because the sheet keeps them too, but are dropped from the filed .xlsx.</p>')
    return "".join(h)


def render_column_table(result: dict) -> str:
    """Which columns actually differ - the <<ColumnTable>> token."""
    if result["status"] != "ok":
        return ""

    rows = []
    for p in result["properties"]:
        if p["status"] != "ok":
            continue
        for kind, label in (("rr4", "RR4"), ("tm30", "TM30")):
            for key, n in sorted(p[kind]["cols"].items(), key=lambda kv: -kv[1]):
                rows.append((p["short"], label, key, n, ""))
            for key, n in sorted(p[kind]["drift_cols"].items(), key=lambda kv: -kv[1]):
                rows.append((p["short"], label, key, n, _KNOWN_DRIFT.get(key, "known drift")))
    if not rows:
        return ('<p style="font-size:13px;color:#166534;margin:0">'
                '✅ Every column of every paired row matches the sheet</p>')

    h = [f'<table style="border-collapse:collapse;width:100%"><tr>'
         f'<th style="{_TH}">Property</th><th style="{_TH}">Register</th><th style="{_TH}">Column</th>'
         f'<th style="{_TH}">Rows</th><th style="{_TH}">Note</th></tr>']
    for short, label, key, n, note in rows:
        style = _MUTED if note else "color:#92400e;font-weight:700;"
        h.append(f'<tr><td style="{_TD}white-space:nowrap">{short}</td>'
                 f'<td style="{_TD}{_MUTED}">{label}</td>'
                 f'<td style="{_TD}{style}">{key}</td>'
                 f'<td style="{_TD}{style}">{n}</td>'
                 f'<td style="{_TD}font-size:11px;color:#64748b">{note or "needs review"}</td></tr>')
    h.append("</table>")
    return "".join(h)


def render_sample_table(result: dict) -> str:
    """Up to three real differing rows per register - the <<SampleTable>>
    token. "nationality differs on 2 rows" is a number to worry about;
    "Nikolaos Pantotis: GRC vs GRL" is something someone can act on this
    morning."""
    if result["status"] != "ok":
        return ""

    rows = []
    for p in result["properties"]:
        if p["status"] != "ok":
            continue
        for kind, label in (("rr4", "RR4"), ("tm30", "TM30")):
            for who, diffs in p[kind]["samples"]:
                rows.append((p["short"], label, who, diffs))
    if not rows:
        return ""

    h = [f'<table style="border-collapse:collapse;width:100%"><tr>'
         f'<th style="{_TH}">Property</th><th style="{_TH}">Register</th>'
         f'<th style="{_TH}">Guest</th><th style="{_TH}">Column</th>'
         f'<th style="{_TH}">Ours</th><th style="{_TH}">Sheet</th></tr>']
    for short, label, who, diffs in rows:
        for i, (key, ov, sv) in enumerate(diffs):
            first = i == 0
            h.append("<tr>")
            if first:
                span = f' rowspan="{len(diffs)}"'
                h.append(f'<td style="{_TD}white-space:nowrap"{span}>{short}</td>'
                         f'<td style="{_TD}{_MUTED}"{span}>{label}</td>'
                         f'<td style="{_TD}"{span}>{_esc(who)}</td>')
            h.append(f'<td style="{_TD}color:#92400e">{key}</td>'
                     f'<td style="{_TD}font-weight:700">{_esc(ov) or "—"}</td>'
                     f'<td style="{_TD}{_MUTED}">{_esc(sv) or "—"}</td></tr>')
    h.append("</table>")
    h.append('<p style="font-size:11px;color:#94a3b8;margin:6px 0 0">'
             'Shows up to 3 rows per register per property, and up to 4 columns per row</p>')
    return "".join(h)


def render_window_table(result: dict) -> str:
    """Sheet export window vs our configured one - the <<WindowTable>> token.
    See _windows() for why this is worth looking at every day."""
    h = [f'<table style="border-collapse:collapse;width:100%"><tr>'
         f'<th style="{_TH}">Property</th><th style="{_TH}">RR4 — Sheet</th>'
         f'<th style="{_TH}">RR4 — Ours</th><th style="{_TH}">TM30 — Sheet</th>'
         f'<th style="{_TH}">TM30 — Ours</th></tr>']
    for p in result["properties"]:
        ok = p["sheet_rr4_window"] == p["our_window"] and p["sheet_rr4_window"]
        tm_ok = p["sheet_tm30_window"] == p["our_tm30_window"] and p["sheet_tm30_window"]
        h.append(f'<tr><td style="{_TD}white-space:nowrap">{p["short"]}</td>'
                 f'<td style="{_TD}">{p["sheet_rr4_window"] or "—"}</td>'
                 f'<td style="{_TD}{"color:#475569;" if ok else _BAD}">{p["our_window"] or "—"}</td>'
                 f'<td style="{_TD}">{p["sheet_tm30_window"] or "—"}</td>'
                 f'<td style="{_TD}{"color:#475569;" if tm_ok else _BAD}">'
                 f'{p["our_tm30_window"] or "—"}</td></tr>')
    h.append("</table>")
    h.append('<p style="font-size:11px;color:#94a3b8;margin:6px 0 0">'
             'Both columns should match their sheet — the sheets change these windows without warning, '
             'and a stale value silently undercounts the register. '
             'TM30 has its own setting per property, separate from RR4\'s.</p>')
    return "".join(h)


def render_tokens(result: dict) -> dict:
    """Everything the email template can substitute."""
    t = result.get("totals") or {"rr4": {}, "tm30": {}}
    day = datetime.strptime(result["date"], "%Y-%m-%d") if result.get("date") else None
    return {
        "Date": day.strftime("%d/%m/%Y") if day else "—",
        "PropertyCount": str(result.get("compared", 0)),
        "Rr4Diff": str(t["rr4"].get("diff_rows", "—")),
        "Tm30Diff": str(t["tm30"].get("diff_rows", "—")),
        "Rr4Rows": f"{t['rr4'].get('ours', '—')} / {t['rr4'].get('sheet', '—')}",
        "Tm30Rows": f"{t['tm30'].get('ours', '—')} / {t['tm30'].get('sheet', '—')}",
        "SummaryTable": render_summary_table(result),
        "ColumnTable": render_column_table(result),
        "SampleTable": render_sample_table(result),
        "WindowTable": render_window_table(result),
    }


def subject_summary(result: dict) -> str:
    if result["status"] != "ok":
        return "not comparable yet"
    t = result["totals"]
    bad = t["rr4"]["diff_rows"] + t["tm30"]["diff_rows"] \
        + t["rr4"]["only_ours"] + t["rr4"]["only_sheet"] \
        + t["tm30"]["only_ours"] + t["tm30"]["only_sheet"]
    return "matches sheet completely" if bad == 0 else f"{bad} rows need review"
